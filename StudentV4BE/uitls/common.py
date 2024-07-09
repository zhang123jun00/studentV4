import openpyxl, uuid, hashlib


def read_excel_import(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['student']
    students = []
    keys = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address']
    for row in sheet.rows:
        temp_dict = {}
        for index, cell in enumerate(row):
            temp_dict[keys[index]] = cell.value
        students.append(temp_dict)
    return students


def get_random_str():
    uuid_var = uuid.uuid4()
    uuid_str = str(uuid_var).encode('utf-8')
    md5 = hashlib.md5()
    md5.update(uuid_str)
    return md5.hexdigest()


def write_to_excel(data: list, path: str):
    """把数据写入到excle"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'student'
    keys = data[0].keys()
    for index, item in enumerate(data):
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))
    workbook.save(path)


